import json
from pathlib import Path
from kslingo.utils.fs import ensure_dir
import re
import csv
from kslingo.utils.fs import validate_file
from kslingo.utils.text import normalize_markdown_title
from kslingo.utils.text import normalize_separator
from kslingo.utils.text import extract_markdown_metadata
from kslingo.utils.text import remove_between
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from collections import OrderedDict
from kslingo.parsers.json import write_json
from kslingo.parsers.markdown import get_phrases_markdown
import markdown
from weasyprint import HTML
from kslingo.support import get_supported_languages
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def Convert_json2md(json_path: str, md_output_path: str, learn_lang: str, native_lang: str) -> None:
    """
    Converts a JSON file into Markdown format with flags (%%A2,W,E%%).
    If enabled=True, the phrase in the learning language is bolded.
    If isword=True, the flag W is used; otherwise, the flag P is used.
    """
    print("start Convert_json2md")

    json_path = Path(json_path)
    md_output_path = Path(md_output_path)

    if not json_path.is_file():
        raise FileNotFoundError(f"Input file '{json_path}' doesn't exist")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    lines = []

    for section in data:
        category = section.get("category", {})

        # get learn language
        cat_learn = category.get(learn_lang)
        if not cat_learn:
            cat_learn = category.get("en", "").strip()
        cat_learn = cat_learn or "Unknown"

        # get native language
        cat_native = category.get(native_lang, "").strip() or "Unknown"

        lines.append(f"### {cat_learn} - {cat_native}")

        for phrase in section.get("phrases", []):
            level = phrase.get("level", "A1")
            enabled = phrase.get("enabled", False)
            isword = phrase.get("isword", False)
            translations = phrase.get("translations")

            # skip if doesn't valid
            if not isinstance(translations, dict):
                continue

            left = translations.get(learn_lang) or translations.get("en")
            right = translations.get(native_lang)
            if not left or not right:
                continue

            left = left.strip()
            right = right.strip()
            if not left or not right:
                continue

            flag_w_p = "W" if isword else "P"
            flag_e_d = "E" if enabled else "D"
            flags = f"%%{level},{flag_w_p},{flag_e_d}%%"

            # if enabled -> bold
            if enabled:
                line = f"- {flags} **{left}** - {right}"
            else:
                line = f"- {flags} {left} - {right}"

            lines.append(line)

        lines.append("")  # empty line between sections

    # Create dir if doesn't exist
    md_output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(md_output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"Markdown file created: {md_output_path}")

# def add_prefix_on_markdown(md_input_path: str, md_output_path: str, prefix: str = "%%A2,W,D%%") -> None:
#     """
#     Adds a given prefix to all phrase lines in a Markdown file, 
#     excluding category headers (lines starting with '###') and empty lines.

#     This is useful for tagging phrases with metadata (e.g., level, type, status)
#     in preparation for conversion to structured formats like JSON.

#     Args:
#         md_input_path (str): Path to the original Markdown (.md) file.
#         md_output_path (str): Path where the modified Markdown file will be saved.
#         prefix (str): The prefix string to prepend to each phrase line.
#     """
    
#     # TODO: za sada bezuslovno ubacuje prefix, napravi da ubacuje samo ako ga nema.
    
#     input_path = Path(md_input_path)
#     output_path = Path(md_output_path)

#     with open(input_path, "r", encoding="utf-8") as f:
#         lines = f.readlines()

#     updated_lines = []
#     for line in lines:
#         stripped = line.strip()

#         if not stripped:
#             updated_lines.append("")
#         elif stripped.startswith("###"):
#             updated_lines.append(line.rstrip())
#         else:
#             updated_lines.append(f"{prefix} {line.strip()}")

#     output_path.parent.mkdir(parents=True, exist_ok=True)
#     with open(output_path, "w", encoding="utf-8") as f:
#         f.write("\n".join(updated_lines) + "\n")
        
#     print(f"Created prefixed markdown file : {output_path}")
    

def Convert_md2json(
    md_input_path: str,
    json_output_path: str,
    learn_lang: str,
    native_lang: str,
) -> None:
    """
    Converts parsed Markdown phrases (from get_phrases_markdown) into structured JSON.
    """

    print("start Convert_md2json")

    phrases = get_phrases_markdown(md_input_path, learn_lang, native_lang)

    if not phrases:
        print("ERROR: phrases is empty array!")
        return

    # Debug print
    for i, sec in enumerate(phrases):
        print(f"Section {i}: {sec['title']}")
        for j, phrase in enumerate(sec["phrases"]):
            print(f"  [{i}][{j}] {phrase}")

    sections = []
    supported_langs = get_supported_languages()

    for sec in phrases:
        # --- CATEGORY (section title) ---
        title_text = normalize_separator(sec["title"])
        parts = title_text.split(" - ", maxsplit=1)

        # Uvek napravi category sa svim jezicima
        category = {}
        for lang in supported_langs:
            if len(parts) == 2:
                if lang == learn_lang:
                    category[lang] = parts[0].strip()
                elif lang == native_lang:
                    category[lang] = parts[1].strip()
                else:
                    category[lang] = ""
            else:
                # Ako nema separatora " - "
                if lang in (learn_lang, native_lang):
                    category[lang] = title_text.strip()
                else:
                    category[lang] = ""

        current_phrases = []

        for phrase in sec["phrases"]:
            # Primer: {'flags': ['A2', 'W', 'E'], 'only_learn': False, 'hu': 'lazy', 'sr': 'lenj'}

            flags = phrase.get("flags", [])
            level = flags[0] if len(flags) > 0 else ""
            isword = "W" in flags
            enabled = "E" in flags

            # --- TRANSLATIONS: popuni sve jezike iz supported liste ---
            translations = {}
            for lang in supported_langs:
                if lang == learn_lang or lang == native_lang:
                    translations[lang] = phrase.get(lang, "")
                else:
                    translations[lang] = ""

            current_phrases.append({
                "level": level,
                "enabled": enabled,
                "isword": isword,
                "translations": translations
            })

        if current_phrases:
            sections.append({
                "category": category,
                "phrases": current_phrases
            })

    ensure_dir(str(Path(json_output_path).parent))
    write_json(sections, json_output_path)
    print(f"Created JSON file : {json_output_path}")


# def Convert_json2csv(json_input_path: str, csv_output_path: str) -> None:
#     """
#     Loads a JSON file and converts it to a CSV file in a flat format.

#     Each phrase becomes a row. Categories are visually separated with an empty row.

#     Args:
#         json_input_path (str): Path to the input JSON file.
#         csv_output_path (str): Path to save the generated CSV file.
#     """
    
#     print("start Convert_json2csv")
    
#     json_input_path = Path(json_input_path)
#     csv_output_path = Path(csv_output_path)

#     if not json_input_path.exists():
#         raise FileNotFoundError(f"JSON file does not exist: {json_input_path}")

#     with open(json_input_path, "r", encoding="utf-8") as f:
#         data = json.load(f)

#     with open(csv_output_path, "w", newline="", encoding="utf-8") as csvfile:
#         writer = csv.writer(csvfile)

#         for category_block in data:
#             category = category_block.get("category", {})
#             phrases = category_block.get("phrases", [])

#             # Write category as a row with translations, no metadata
#             writer.writerow([
#                 "", "", "",  # empty: level, isword, enabled
#                 category.get("hu", ""),
#                 category.get("sr", ""),
#                 category.get("en", "")
#             ])

#             # Write all phrases for the category
#             for phrase in phrases:
#                 writer.writerow([
#                     phrase.get("level", ""),
#                     "Yes" if phrase.get("isword", False) else "No",
#                     "Yes" if phrase.get("enabled", False) else "No",
#                     phrase["translations"].get("hu", ""),
#                     phrase["translations"].get("sr", ""),
#                     phrase["translations"].get("en", ""),
#                 ])

#             # Blank row to separate categories
#             writer.writerow([])
            

def Convert_json2xlsx(json_path: str, xlsx_path: str) -> None:
    print("start Convert_json2xlsx")

    supported_langs = get_supported_languages()

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active

    category_font = Font(size=13, bold=True)
    phrase_font = Font(size=12, bold=False)

    category_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")

    gray_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0")
    )

    # Header row
    header_row = ["Enabled", "Level", "IsWord"] + [lang.upper() for lang in supported_langs]
    ws.append(header_row)
    ws.freeze_panes = "A2"

    # Set fixed width on language columns
    for i, lang in enumerate(supported_langs, start=4):  # Columns D, E, ...
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 30

    # Center align header
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")

    for block in data:
        category = block["category"]
        phrases = block["phrases"]

        # CATEGORY ROW
        category_row = ["", "", ""] + [category.get(lang, "") for lang in supported_langs]
        ws.append(category_row)

        for cell in ws[ws.max_row]:
            cell.font = category_font
            cell.fill = category_fill
            cell.border = gray_border

        # PHRASES
        for phrase in phrases:
            row = [
                "1" if phrase.get("enabled", False) else "0",
                phrase.get("level", ""),
                "1" if phrase.get("isword", False) else "0",
            ] + [
                phrase["translations"].get(lang, "") for lang in supported_langs
            ]

            ws.append(row)
            current_row = ws[ws.max_row]
            is_disabled = row[0] == "0"

            for i, cell in enumerate(current_row, start=1):
                cell.border = gray_border
                cell.font = phrase_font
                if i in (1, 2, 3):  # Enabled, Level, IsWord
                    cell.alignment = Alignment(horizontal="center")

                cell.fill = red_fill if is_disabled else green_fill

        # Empty row between categories
        ws.append([])

    # Conditional formatting
    max_row = ws.max_row
    last_col = get_column_letter(3 + len(supported_langs))  # Example: "F", "G", ...
    ws.conditional_formatting.add(
        f"A2:{last_col}{max_row}",
        FormulaRule(formula=['AND(ISNUMBER($A2), $A2=1)'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"A2:{last_col}{max_row}",
        FormulaRule(formula=['AND(ISNUMBER($A2), $A2=0)'], fill=red_fill)
    )

    wb.save(xlsx_path)
    print(f"XLSX saved to {xlsx_path}")



    # Dodaj uslovno formatiranje koje reaguje na vrednost A u svakom redu
    max_row = ws.max_row
    ws.conditional_formatting.add(
        f"A2:F{max_row}",
        FormulaRule(formula=['AND(ISNUMBER($A2), $A2=1)'], fill=green_fill)
    )

    ws.conditional_formatting.add(
        f"A2:F{max_row}",
        FormulaRule(formula=['AND(ISNUMBER($A2), $A2=0)'], fill=red_fill)
    )

    # Save to XLSX file
    wb.save(xlsx_path)
    print(f"XLSX saved to {json_path}")
    
    
def Convert_xlsx2json(xlsx_path: str, json_path: str) -> None:
    print("start Convert_xlsx2json")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    data = []
    current_block = None

    for row in rows:
        enabled, level, isword, hu, sr, en = row

        # Empty row — skip
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        # Row with no flags but has translations = CATEGORY
        if enabled in (None, "", 0) and level in (None, "", 0) and isword in (None, "", 0):
            # New category block
            current_block = {
                "category": {
                    "hu": hu or "",
                    "sr": sr or "",
                    "en": en or ""
                },
                "phrases": []
            }
            data.append(current_block)
            continue

        # Otherwise it's a phrase row
        phrase = OrderedDict([
            ("level", str(level).strip()),
            ("enabled", str(enabled).strip() == "1"),
            ("isword", str(isword).strip() == "1"),
            ("translations", {
                "hu": hu or "",
                "sr": sr or "",
                "en": en or ""
            })
        ])

        if current_block is not None:
            current_block["phrases"].append(phrase)
        else:
            print("Warning: Phrase row found before any category, skipping")

    # Save to JSON file
    write_json(data, json_path)

    print(f"JSON saved to {json_path}")

def Generate_pdf_from_md(md_path: str, pdf_path: str) -> str:
    """
    Converts a Markdown (.md) file to PDF and saves it to the specified directory.

    Args:
        md_path (str): Path to the Markdown file.
        pdf_path (str): Path to the PDF file.

    Returns:
        str: Path to the generated PDF file.
    """
    
    print("start Generate_pdf_from_md")

    validate_file(md_path, ".md")

    # load content from MD 
    with open(md_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    # convert MD to HTML
    html_text = markdown.markdown(md_text)

    # convert HTML to PDF
    HTML(string=html_text).write_pdf(pdf_path)

    validate_file(pdf_path, ".pdf")

    return

